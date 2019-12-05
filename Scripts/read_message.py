import pandas as pd
from openpyxl import load_workbook

def read_msg(m_id):
    from apiclient import discovery
    from apiclient import errors
    from httplib2 import Http
    from oauth2client import file, client, tools
    import base64
    from bs4 import BeautifulSoup
    import re
    import time
    import dateutil.parser as parser
    from datetime import datetime
    import datetime
    import csv
    import email

    # Creating a storage.JSON file with authentication details
    SCOPES = 'https://www.googleapis.com/auth/gmail.readonly' # we are using modify and not readonly, as we will be marking the messages Read
    store = file.Storage('storage.json') 
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('credentials.json', SCOPES)
        creds = tools.run_flow(flow, store)
    GMAIL = discovery.build('gmail', 'v1', http=creds.authorize(Http()))

    user_id =  'me'

    message = GMAIL.users().messages().get(userId=user_id, id=m_id).execute() # fetch the message using API
    payld = message['payload'] # get payload of the message 
    headr = payld['headers'] # get header of the payload
    
    try:# Fetching message body
        mssg_parts = payld['parts'] # fetching the message parts
        part_one  = mssg_parts[0] # fetching first element of the part 
        part_body = part_one['body'] # fetching body of the message
        part_data = part_body['data'] # fetching data from the body
        msg_str = base64.urlsafe_b64decode(part_data.encode('UTF8'))
        mssg_body = email.message_from_bytes(msg_str)
        return mssg_body    
    except:
        return ""

def main_func(df):
    message_list = []
    for i in df:
        message_dict = {}
        message_dict['MID'] = i
        msg = read_msg(i)
        print(msg)
        message_dict['Message'] = msg
        try:
            message_list.append(message_dict)
        except:
            pass

    db = pd.DataFrame(message_list)
    print(db.head())
    book = load_workbook('text.xlsx')
    writer = pd.ExcelWriter('text.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    print(writer.sheets['Sheet1'].max_row)
    db.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = True,header= False)
    writer.save()
        
if __name__ == "__main__":
    main_func(['166f98fe819af709','16701f2471af952a'])
    
   

    

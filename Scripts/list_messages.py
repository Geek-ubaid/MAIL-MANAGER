def list_message():
    from apiclient import discovery
    from apiclient import errors
    from httplib2 import Http
    from oauth2client import file, client, tools
    import base64
    from bs4 import BeautifulSoup
    import re
    import time
    import dateutil.parser as parser
    from datetime import datetime
    import datetime
    import csv
    import datetime
    import pandas
    from openpyxl import load_workbook

    # Creating a storage.JSON file with authentication details
    SCOPES = 'https://www.googleapis.com/auth/gmail.readonly' # we are using modify and not readonly, as we will be marking the messages Read
    store = file.Storage('storage.json') 
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('credentials.json', SCOPES)
        creds = tools.run_flow(flow, store)
    GMAIL = discovery.build('gmail', 'v1', http=creds.authorize(Http()))

    user_id =  'me'

    unread_msgs = GMAIL.users().messages().list(userId='me',maxResults=2).execute()

    # We get a dictonary. Now reading values for the key 'messages'
    mssg_list = unread_msgs['messages']

    print ("Total messages", str(len(mssg_list)))

    final_list = [ ]

    for mssg in mssg_list:
        
        temp_dict = { }
        m_id = mssg['id'] # get id of individual message
        message = GMAIL.users().messages().get(userId=user_id, id=m_id).execute() # fetch the message using API
        temp_dict['Label'] = " ".join(message['labelIds'])
        temp_dict['Size'] = message['sizeEstimate']
        payld = message['payload'] # get payload of the message 
        headr = payld['headers'] # get header of the payload

        temp_dict['MID'] = m_id
        
        
        for one in headr: # getting the Subject
            if one['name'] == 'Subject':
                msg_subject = one['value']
                temp_dict['Subject'] = msg_subject
            else:
                pass
        

        for two in headr: # getting the date
            if two['name'] == 'Date':
                msg_date = two['value']
                date_parse = (parser.parse(msg_date))
                m_date = (date_parse.date())
                temp_dict['Date'] = str(m_date)
                temp_dict['Time'] = "{}:{}:{}".format(str(date_parse.hour),str(date_parse.minute),str(date_parse.second))
            else:
                pass

        for three in headr: # getting the Sender
            if three['name'] == 'From':
                msg_from = three['value']
                temp_dict['Sender'] = msg_from
            else:
                pass
        for four in headr:
            if four['name'] == 'To':
                temp_dict['To'] = four['value']

        temp_dict['Snippet'] = message['snippet'][:len(message['snippet'])] # fetching message snippet
        try:
            parts = message['payload']['parts']
            temp_dict['Attach_file'] = ""
            temp_dict['Attach_size'] = ""
            temp_dict['Attach_ID'] = ""
            temp_dict['Attach_Name'] = ""
            for part in message['payload']['parts']:
                newvar = part['body']
                if 'attachmentId' in newvar:
    
                    temp_dict['Attachment'] = 'Yes'
                    temp_dict['Attach_file'] = part['mimeType']
                    temp_dict['Attach_Name'] = part['filename']
                    temp_dict['Attach_size'] = newvar['size']
                    for hed in part['headers']:
                        if hed['name'] == 'X-Attachment-Id':
                            temp_dict['Attach_ID'] = hed['value']
                    
                else:
                    temp_dict['Attachment'] = 'No'
        except:
            temp_dict['Attachment'] = 'No'
         
        
        try:
       
            final_list.append(temp_dict)

        except:
            pass
        
        # This will mark the messagea as read
    ##	GMAIL.users().messages().modify(userId=user_id, id=m_id,body={ 'removeLabelIds': ['UNREAD']}).execute() 
            
    print ("Total messaged retrived: ", str(len(final_list)))
    db = pandas.DataFrame(final_list)
    #exporting the values as .csv
    book = load_workbook('message.xlsx')
    writer = pandas.ExcelWriter('message.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    db.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = True,header= False)
    writer.save()
    
if __name__ == "__main__":
    list_message()
    
